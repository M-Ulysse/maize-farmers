import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

EXCEL_FILE = "Maize_Farmers_Data.xlsx"

COLUMNS = ["Farmer ID", "Farmer Name", "District", "Product",
           "Plot Size (acres)", "Quantity (kg)", "Price per kg (RWF)", "Total Revenue (RWF)"]

DISTRICTS = [
    "Bugesera", "Gatsibo", "Gasabo", "Gicumbi", "Huye", "Kamonyi",
    "Karongi", "Kayonza", "Kirehe", "Muhanga", "Musanze", "Nyagatare",
    "Nyamagabe", "Nyanza", "Nyaruguru", "Rubavu", "Rulindo", "Rusizi",
    "Rutsiro", "Rwamagana"
]

GREEN_DARK   = "#1B5E20"
GREEN_MID    = "#2E7D32"
GREEN_LIGHT  = "#4CAF50"
GREEN_PALE   = "#E8F5E9"
WHITE        = "#FFFFFF"
CHART_COLORS = ["#1B5E20", "#2E7D32", "#388E3C", "#43A047", "#4CAF50",
                "#66BB6A", "#81C784", "#A5D6A7", "#C8E6C9", "#E8F5E9"]


def load_data():
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame(columns=COLUMNS)
    df = pd.read_excel(EXCEL_FILE, sheet_name="Maize Farmers Data")
    df = df[df["Farmer ID"].notna()]
    df = df[df["Farmer ID"] != "TOTAL"]
    return df.reset_index(drop=True)


def save_data(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maize Farmers Data"
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    alt_fill    = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    center      = Alignment(horizontal="center", vertical="center")
    border      = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"),  bottom=Side(style="thin"))
    for col, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[1].height = 25
    for row_idx, row in df.iterrows():
        fill = alt_fill if (row_idx + 2) % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=row[col_name])
            cell.border = border; cell.alignment = center
            if fill.fill_type: cell.fill = fill
    total_row = len(df) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=1).border = border
    for col in range(2, 9):
        cell = ws.cell(row=total_row, column=col)
        cell.border = border
        if col in (6, 8):
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            cell.font = Font(bold=True); cell.alignment = center
    col_widths = [10, 28, 16, 10, 18, 16, 22, 22]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)


# ── Global CSS ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Maize Farmers Dashboard", page_icon="🌽", layout="wide")

st.markdown(f"""
<style>
  /* Background */
  .stApp {{ background-color: {WHITE}; }}
  [data-testid="stSidebar"] {{ background-color: {GREEN_DARK} !important; }}
  [data-testid="stSidebar"] * {{ color: {WHITE} !important; }}
  [data-testid="stSidebar"] .stSelectbox label,
  [data-testid="stSidebar"] .stMultiSelect label,
  [data-testid="stSidebar"] .stNumberInput label {{ color: {WHITE} !important; }}

  /* Header banner */
  .dashboard-header {{
      background: linear-gradient(135deg, {GREEN_DARK} 0%, {GREEN_LIGHT} 100%);
      padding: 28px 36px; border-radius: 12px; margin-bottom: 24px;
      color: {WHITE};
  }}
  .dashboard-header h1 {{ color: {WHITE}; font-size: 2.2rem; margin: 0; }}
  .dashboard-header p  {{ color: #C8E6C9; margin: 4px 0 0; font-size: 1rem; }}

  /* Metric cards */
  [data-testid="stMetric"] {{
      background: {GREEN_PALE}; border-left: 5px solid {GREEN_MID};
      border-radius: 10px; padding: 14px 18px;
  }}
  [data-testid="stMetricLabel"] {{ color: {GREEN_DARK} !important; font-weight: 600; }}
  [data-testid="stMetricValue"] {{ color: {GREEN_DARK} !important; font-size: 1.6rem !important; }}

  /* Section headings */
  .section-title {{
      color: {GREEN_DARK}; font-size: 1.2rem; font-weight: 700;
      border-bottom: 3px solid {GREEN_LIGHT}; padding-bottom: 6px; margin: 24px 0 12px;
  }}

  /* Buttons */
  .stButton > button {{
      background-color: {GREEN_MID} !important; color: {WHITE} !important;
      border: none !important; border-radius: 8px !important;
      font-weight: 600 !important;
  }}
  .stButton > button:hover {{ background-color: {GREEN_DARK} !important; }}

  /* Form */
  [data-testid="stForm"] {{
      background: {GREEN_PALE}; border-radius: 10px;
      border: 1px solid #C8E6C9; padding: 16px;
  }}

  /* Download button */
  .stDownloadButton > button {{
      background-color: {GREEN_DARK} !important; color: {WHITE} !important;
      border-radius: 8px !important; font-weight: 600 !important;
  }}

  /* Dataframe header */
  [data-testid="stDataFrame"] thead {{ background-color: {GREEN_MID} !important; color: {WHITE} !important; }}

  /* Divider */
  hr {{ border-color: {GREEN_PALE}; }}
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="dashboard-header">
  <h1>🌽 Maize Farmers Dashboard</h1>
  <p>Rwanda — Track farmers, production, pricing, and revenue at a glance</p>
</div>
""", unsafe_allow_html=True)

df = load_data()

# ── Sidebar ───────────────────────────────────────────────────────────────────
st.sidebar.markdown("## Filters")
district_filter = st.sidebar.multiselect(
    "District",
    sorted(df["District"].dropna().unique()) if not df.empty else DISTRICTS
)
min_plot = st.sidebar.number_input("Min Plot Size (acres)", min_value=0.0, value=0.0, step=0.5)
max_plot = st.sidebar.number_input("Max Plot Size (acres)", min_value=0.0, value=100.0, step=0.5)

filtered = df.copy()
if district_filter:
    filtered = filtered[filtered["District"].isin(district_filter)]
filtered = filtered[
    (filtered["Plot Size (acres)"] >= min_plot) &
    (filtered["Plot Size (acres)"] <= max_plot)
]

# ── Metrics ───────────────────────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
c1.metric("Total Farmers",       f"{len(filtered)}")
c2.metric("Total Quantity (kg)", f"{filtered['Quantity (kg)'].sum():,.0f}"       if not filtered.empty else "0")
c3.metric("Total Revenue (RWF)", f"{filtered['Total Revenue (RWF)'].sum():,.0f}" if not filtered.empty else "0")
c4.metric("Avg Plot Size (acres)",f"{filtered['Plot Size (acres)'].mean():.2f}"  if not filtered.empty else "0.00")

# ── Charts ────────────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">Analytics</div>', unsafe_allow_html=True)

chart_col1, chart_col2, chart_col3 = st.columns(3)

# Chart 1 — Price vs Total Revenue (scatter)
with chart_col1:
    st.markdown("**Price vs Total Revenue**")
    if not filtered.empty:
        fig1 = px.scatter(
            filtered,
            x="Price per kg (RWF)",
            y="Total Revenue (RWF)",
            size="Quantity (kg)",
            color="District",
            hover_name="Farmer Name",
            color_discrete_sequence=CHART_COLORS,
            template="plotly_white",
        )
        fig1.update_layout(
            plot_bgcolor=WHITE,
            paper_bgcolor=WHITE,
            font_color=GREEN_DARK,
            title_font_color=GREEN_DARK,
            legend_title_text="District",
            margin=dict(l=10, r=10, t=30, b=10),
            height=360,
        )
        fig1.update_xaxes(showgrid=True, gridcolor=GREEN_PALE)
        fig1.update_yaxes(showgrid=True, gridcolor=GREEN_PALE)
        st.plotly_chart(fig1, use_container_width=True)
    else:
        st.info("No data to display.")

# Chart 2 — District vs Total Quantity (bar)
with chart_col2:
    st.markdown("**District vs Quantity (kg)**")
    if not filtered.empty:
        dist_qty = filtered.groupby("District")["Quantity (kg)"].sum().reset_index().sort_values("Quantity (kg)", ascending=False)
        fig2 = px.bar(
            dist_qty,
            x="District",
            y="Quantity (kg)",
            color="Quantity (kg)",
            color_continuous_scale=[[0, GREEN_PALE], [0.5, GREEN_MID], [1, GREEN_DARK]],
            template="plotly_white",
        )
        fig2.update_layout(
            plot_bgcolor=WHITE,
            paper_bgcolor=WHITE,
            font_color=GREEN_DARK,
            coloraxis_showscale=False,
            margin=dict(l=10, r=10, t=30, b=10),
            height=360,
        )
        fig2.update_xaxes(tickangle=-35, showgrid=False)
        fig2.update_yaxes(showgrid=True, gridcolor=GREEN_PALE)
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.info("No data to display.")

# Chart 3 — Plot Size vs Quantity (scatter with trendline)
with chart_col3:
    st.markdown("**Plot Size vs Quantity**")
    if not filtered.empty:
        fig3 = px.scatter(
            filtered,
            x="Plot Size (acres)",
            y="Quantity (kg)",
            color="District",
            trendline="ols",
            hover_name="Farmer Name",
            color_discrete_sequence=CHART_COLORS,
            template="plotly_white",
        )
        fig3.update_traces(marker=dict(size=10))
        fig3.update_layout(
            plot_bgcolor=WHITE,
            paper_bgcolor=WHITE,
            font_color=GREEN_DARK,
            legend_title_text="District",
            margin=dict(l=10, r=10, t=30, b=10),
            height=360,
        )
        fig3.update_xaxes(showgrid=True, gridcolor=GREEN_PALE)
        fig3.update_yaxes(showgrid=True, gridcolor=GREEN_PALE)
        st.plotly_chart(fig3, use_container_width=True)
    else:
        st.info("No data to display.")

# ── Data table ────────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">Farmer Records</div>', unsafe_allow_html=True)
st.dataframe(filtered, use_container_width=True, hide_index=True)

# ── Add farmer ────────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">Add New Farmer</div>', unsafe_allow_html=True)
with st.form("add_farmer_form", clear_on_submit=True):
    a1, a2, a3 = st.columns(3)
    farmer_id   = a1.text_input("Farmer ID", placeholder="e.g. F016")
    farmer_name = a2.text_input("Farmer Name")
    district    = a3.selectbox("District", DISTRICTS)

    a4, a5, a6 = st.columns(3)
    plot_size = a4.number_input("Plot Size (acres)", min_value=0.1, step=0.5, value=1.0)
    quantity  = a5.number_input("Quantity (kg)",     min_value=1,   step=50,  value=500)
    price     = a6.number_input("Price per kg (RWF)",min_value=1,   step=10,  value=350)

    if st.form_submit_button("Add Farmer", type="primary"):
        if not farmer_id or not farmer_name:
            st.error("Farmer ID and Name are required.")
        elif farmer_id in df["Farmer ID"].values:
            st.error(f"Farmer ID '{farmer_id}' already exists.")
        else:
            new_row = {
                "Farmer ID": farmer_id, "Farmer Name": farmer_name,
                "District": district,   "Product": "Maize",
                "Plot Size (acres)": plot_size, "Quantity (kg)": quantity,
                "Price per kg (RWF)": price,
                "Total Revenue (RWF)": quantity * price,
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            save_data(df)
            st.success(f"Farmer '{farmer_name}' added successfully!")
            st.rerun()

# ── Delete farmer ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">Delete a Farmer</div>', unsafe_allow_html=True)
if not df.empty:
    to_delete = st.selectbox("Select farmer", df["Farmer ID"] + " — " + df["Farmer Name"])
    if st.button("Delete Selected Farmer"):
        fid = to_delete.split(" — ")[0]
        df = df[df["Farmer ID"] != fid].reset_index(drop=True)
        save_data(df)
        st.success(f"Farmer {fid} deleted.")
        st.rerun()
else:
    st.info("No farmers to delete.")

# ── Download ──────────────────────────────────────────────────────────────────
st.markdown("---")
if os.path.exists(EXCEL_FILE):
    with open(EXCEL_FILE, "rb") as f:
        st.download_button(
            label="⬇ Download Excel File",
            data=f,
            file_name=EXCEL_FILE,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
